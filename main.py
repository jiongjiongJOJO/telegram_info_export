import os
import re
import pandas as pd
import tkinter as tk
from app.version import VERSION
from ttkthemes import ThemedTk
from openpyxl import load_workbook
from telethon.sync import TelegramClient
from tkinter import ttk, messagebox, simpledialog
from openpyxl.worksheet.table import Table, TableStyleInfo


def get_telegram_info(api_name, api_id, api_hash, phone_number, proxy):

    def get_display_width(s):
        """
        计算字符串的显示宽度
        中文字符算2个宽度，英文字符算1个宽度
        """
        if not isinstance(s, str):
            s = str(s)
        # 使用正则表达式分辨中文字符和非中文字符
        zh_pattern = re.compile(u'[\u4e00-\u9fa5]')
        width = 0
        for ch in s:
            if zh_pattern.match(ch):
                width += 2
            else:
                width += 1
        return width

    # 创建并连接客户端
    print(proxy)
    client = TelegramClient(api_name, int(api_id), api_hash, proxy=proxy)
    client.connect()

    if not client.is_user_authorized():
        client.send_code_request(phone_number)
        verification_code = simpledialog.askstring(title="验证码", prompt="请输入验证码:")
        client.sign_in(phone_number, verification_code)

    # 获取对话
    groups_and_channels = []
    bots = []

    for dialog in client.iter_dialogs():
        if dialog.is_group:
            groups_and_channels.append((dialog.entity, '群组', dialog.name))
        elif dialog.is_channel:
            groups_and_channels.append((dialog.entity, '频道', dialog.name))
        elif dialog.entity.bot:
            bots.append((dialog.entity.username, dialog.entity.first_name, dialog.entity.id))

    # 准备导出数据
    group_channel_data = []
    for entity, chat_type, chat_title in groups_and_channels:
        try:
            username = getattr(entity, 'username', None)
            if username:
                invite_link = f"https://t.me/{username}"
            else:
                invite_link = ""
            group_channel_data.append({
                "类型": chat_type,
                "ID": entity.id,
                "名称": chat_title,
                "邀请链接": invite_link
            })
        except Exception:
            group_channel_data.append({
                "类型": chat_type,
                "ID": entity.id,
                "名称": chat_title,
                "邀请链接": ""
            })

    bot_data = []
    for username, first_name, user_id in bots:
        bot_data.append({
            "用户名": f"@{username}",
            "昵称": first_name,
            "用户ID": user_id
        })

    # 创建DataFrame
    df_groups_channels = pd.DataFrame(group_channel_data)
    df_bots = pd.DataFrame(bot_data)

    # 导出到Excel
    file_path = "telegram_info.xlsx"
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_groups_channels.to_excel(writer, sheet_name="群组和频道", index=False)
        df_bots.to_excel(writer, sheet_name="机器人", index=False)

    # 加载Excel文件并添加筛选功能
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 创建表格
        tab = Table(displayName=f"Table_{sheet_name}", ref=ws.dimensions)
        # 添加样式
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    length = get_display_width(cell.value)
                    if length > max_length:
                        max_length = length
                except:
                    pass
            adjusted_width = (max_length + 4)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    # 断开客户端
    client.disconnect()

    messagebox.showinfo("保存成功", "信息已导出到 {} 文件中".format('telegram_info.xlsx'))


def get_proxy_env():
    # 读取环境变量中的代理设置
    proxy_env = os.getenv('HTTP_PROXY') or os.getenv('ALL_PROXY')
    proxy = None
    re_pattern = r'''
    ^(?P<protocol>socks5|socks4|http|https)://(?:(?P<user>[^:]+)(?::(?P<password>[^@]+))?@)?(?P<host>[^:]+):(?P<port>\d+)$
    '''.strip()
    if proxy_env:
        proxy_url_pattern = re.compile(re_pattern)
        match = proxy_url_pattern.match(proxy_env)
        if match:
            proxy_protocol = match.group('protocol')
            proxy_host = match.group('host')
            proxy_port = int(match.group('port'))
            proxy_username = match.group('user')
            proxy_password = match.group('password')
            if proxy_username and proxy_password:
                proxy = (proxy_protocol, proxy_host, proxy_port, proxy_username, proxy_password)
            else:
                proxy = (proxy_protocol, proxy_host, proxy_port)
    return proxy


class Main:
    def __init__(self):
        self.proxy = get_proxy_env()
        # 主界面
        self.root = ThemedTk(theme="yaru")
        self.root.title("获取TG信息")

        ttk.Label(self.root, text="API Name:").grid(row=0, column=0, padx=10, pady=5)
        ttk.Label(self.root, text="API ID:").grid(row=1, column=0, padx=10, pady=5)
        ttk.Label(self.root, text="API Hash:").grid(row=2, column=0, padx=10, pady=5)
        ttk.Label(self.root, text="手机号:").grid(row=3, column=0, padx=10, pady=5)

        api_name = ttk.Entry(self.root, width=36)
        api_id = ttk.Entry(self.root, width=36)
        api_hash = ttk.Entry(self.root, width=36)
        phone_number = ttk.Entry(self.root, width=36)

        api_name.grid(row=0, column=1, padx=10, pady=5)
        api_id.grid(row=1, column=1, padx=10, pady=5)
        api_hash.grid(row=2, column=1, padx=10, pady=5)
        phone_number.grid(row=3, column=1, padx=10, pady=5)

        button_frame = ttk.Frame(self.root)
        button_frame.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(
            button_frame,
            text="代理设置",
            command=self.open_settings_window
        ).pack(side="left", padx=10)

        ttk.Button(
            button_frame,
            text="获取并保存数据",
            command=lambda: get_telegram_info(
                api_name.get(),
                api_id.get(),
                api_hash.get(),
                phone_number.get(),
                self.proxy
            )
        ).pack(side="left", padx=10)

        self.root.mainloop()

    # 设置窗口
    def open_settings_window(self):
        settings_window = tk.Toplevel(self.root)
        settings_window.title("设置")

        ttk.Label(settings_window, text="代理类型:").grid(row=0, column=0, padx=10, pady=5)
        ttk.Label(settings_window, text="代理IP:").grid(row=1, column=0, padx=10, pady=5)
        ttk.Label(settings_window, text="代理端口:").grid(row=2, column=0, padx=10, pady=5)
        ttk.Label(settings_window, text="用户名:").grid(row=3, column=0, padx=10, pady=5)
        ttk.Label(settings_window, text="密码:").grid(row=4, column=0, padx=10, pady=5)

        proxy_type = ttk.Entry(settings_window)
        proxy_ip = ttk.Entry(settings_window)
        proxy_port = ttk.Entry(settings_window)
        username = ttk.Entry(settings_window)
        password = ttk.Entry(settings_window, show="*")

        proxy_type.grid(row=0, column=1, padx=10, pady=5)
        proxy_ip.grid(row=1, column=1, padx=10, pady=5)
        proxy_port.grid(row=2, column=1, padx=10, pady=5)
        username.grid(row=3, column=1, padx=10, pady=5)
        password.grid(row=4, column=1, padx=10, pady=5)

        entry_list = proxy_type, proxy_ip, proxy_port, username, password

        if self.proxy:
            for i in range(len(self.proxy)):
                entry_list[i].insert(0, self.proxy[i])
        ttk.Button(
            settings_window,
            text="保存",
            command=lambda: self.set_proxy_info(
                settings_window, proxy_type, proxy_ip, proxy_port, username, password
            )
        ).grid(row=5, column=0, columnspan=2, pady=10)

    def set_proxy_info(self, settings_window, proxy_type, proxy_ip, proxy_port, username, password):
        proxy_type_str = proxy_type.get()
        proxy_ip_str = proxy_ip.get()
        proxy_port_str = proxy_port.get()
        username_str = username.get()
        password_str = password.get()
        if proxy_type_str and proxy_ip_str and proxy_port_str:
            if username:
                self.proxy = (proxy_type_str, proxy_ip_str, int(proxy_port_str), username_str, password_str)
            else:
                self.proxy = (proxy_type_str, proxy_ip_str, int(proxy_port_str))
        settings_window.destroy()


print("当前程序版本：{}".format(VERSION))
main = Main()
